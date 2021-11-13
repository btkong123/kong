import openpyxl
import os
from openpyxl import load_workbook

def main():
    try:
        f = open('致橡树','r',encoding='utf-8')
        print(f.read)
        f.close()
    except Exception:
        print('open failed')

def open_file():
    fp = open('exercise.xlsx', 'w')
    wb = load_workbook('exercise.xlsx', data_only=True)
    sheet = wb.active
    write_wb = openpyxl.Workbook()
    write_sheet = write_wb.active
    print('开始写入excel')
    #表头信息
    excel_title = [
        'title_1',
        'title_2',
        'title_3'
    ]
    write_sheet.append(excel_title)
    write_wb.save('new.xlsx')

if __name__ == '__main__':
    open_file()